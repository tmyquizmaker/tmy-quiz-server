"""
===========================================
TMY Quiz Maker
Version 5.3

app_controller.py

Contrôle navigation + gestion des quiz locaux + Lobby Multijoueur
===========================================
"""

import threading
from datetime import datetime
from tkinter import filedialog, messagebox

import customtkinter as ctk

from audio.music import resume_music, stop_music
from core.network import NetworkClient
from data.quiz_storage import save_quiz
from library_hub import LibraryHubWindow

from ui.home import HomePage
from ui.join_page import JoinRoomPage
from ui.leaderboard_overlay import TeacherFullDashboard
from ui.levels_menu import LevelsMenuPage
from ui.loading import LoadingScreen
from ui.my_quizzes import MyQuizzesPage
from ui.party_create import PartyCreatePage
from ui.party_game import PartyGamePage
from ui.play_quiz import PlayQuizPage
from ui.play_quiz_manuel import PlayQuizManuelPage  # 👈 CÔTÉ ÉLÈVE
from ui.quiz_loading import QuizLoadingPage
from ui.quiz_lobby import QuizLobbyPage
from ui.result import ResultPage
from ui.tmy_generator import TMYGeneratorPage

from ui.auth_page import AuthPage
from ui.forgot_password_page import ForgotPasswordPage
from ui.manual_quiz_choice_page import ManualQuizChoicePage
from ui.verify_email_page import VerifyEmailPage
from ui.settings_page import SettingsPage
from ui.account_page import AccountPage
from auth_client import session

try:
    from ui.manual_quiz import ManualQuizPage
except ImportError:
    ManualQuizPage = None


class AppController:

    def __init__(self, root):
        self.root = root

        self.network = NetworkClient()  # Initialisation du client réseau WebSocket

        self.current_quiz_settings = {
            "sujet": "",
            "niveau": "",
            "nombre": 0
        }

        self.current_quiz = []
        self.loading_active = False

        # Métadonnées pour la session réseau (Élève / Prof)
        self.current_student_name = "Élève"
        self.current_quiz_title = "Quiz en direct"
        self.current_teacher_name = "Professeur"
        self.current_active_pin = ""  # PIN actif conservé pour l'hôte ou l'élève

        # Dictionnaire pour stocker les salons ouverts
        self.active_lobbies = {}

    # =====================================
    # Démarrage
    # =====================================
    def start(self):
        self.show_loading()

    # =====================================
    # Loading application
    # =====================================
    def show_loading(self):
        self.clear_page()
        self.loading_active = True

        self.loading = LoadingScreen(self.root)
        self.loading.pack(fill="both", expand=True)

        self.run_loading_animation()

    def run_loading_animation(self):
        etapes = self.loading.animation()

        def next_step(index=0):
            if not self.loading_active:
                return

            if index < len(etapes):
                valeur, message = etapes[index]
                self.loading.update_progress(valeur, message)
                self.root.after(500, lambda: next_step(index + 1))
            else:
                self.show_home()

        next_step()

    # =====================================
    # Accueil
    # =====================================
    def show_home(self):
        self.loading_active = False
        self.clear_page()

        self.home = HomePage(
            master=self.root,
            tmy_callback=self.show_tmy_generator,
            manual_callback=self.show_manual_choice,
            quizzes_callback=self.show_my_quizzes,
            multiplayer_callback=self.show_multiplayer,
            join_callback=self.show_join_room,
            settings_callback=self.show_settings,
            account_callback=self.show_account
        )

        self.home.pack(fill="both", expand=True)

    # =====================================
    # Authentification (page complète, pas un popup)
    # =====================================
    def show_auth_page(self, on_success=None):
        """Affiche la page de connexion/inscription. Une fois connecté, exécute
        on_success (l'action que l'utilisateur voulait faire à l'origine), ou
        revient à l'accueil si aucune action n'était en attente."""
        self.clear_page()
        self.auth_page = AuthPage(
            self.root,
            on_success=lambda: self._apres_connexion(on_success),
            back_callback=self.show_home,
            forgot_password_callback=self.show_forgot_password,
            verify_email_callback=self.show_verify_email,
        )
        self.auth_page.pack(fill="both", expand=True)

    def show_forgot_password(self):
        """Page dédiée (pas un popup) pour réinitialiser le mot de passe par code
        envoyé par email. Retourne à la page de connexion une fois terminé."""
        self.clear_page()
        self.forgot_password_page = ForgotPasswordPage(
            self.root,
            back_callback=self.show_auth_page,
            on_success=self.show_auth_page,
        )
        self.forgot_password_page.pack(fill="both", expand=True)

    def show_verify_email(self, email):
        """Page de saisie du code de vérification reçu par email (après
        inscription, ou depuis l'écran de connexion si le compte n'est pas
        encore vérifié). Retourne à la connexion une fois vérifié."""
        self.clear_page()
        self.verify_email_page = VerifyEmailPage(
            self.root,
            email=email,
            back_callback=self.show_auth_page,
            on_success=self.show_auth_page,
        )
        self.verify_email_page.pack(fill="both", expand=True)

    def _apres_connexion(self, on_success):
        if on_success:
            on_success()
        else:
            self.show_home()

    def _requires_login(self, callback):
        """Exécute callback directement si connecté, sinon ouvre la page d'authentification."""
        if session.est_connecte():
            callback()
        else:
            self.show_auth_page(on_success=callback)

    # --- Callbacks complémentaires ---
    def show_multiplayer(self):
        """Vérifie la connexion avant d'ouvrir l'écran de sélection des niveaux (mode en ligne)."""
        self._requires_login(self._show_multiplayer_actual)

    def _show_multiplayer_actual(self):
        """Ouvre l'écran de sélection des 10 niveaux du mode multijoueur."""
        self.clear_page()
        self.levels_page = LevelsMenuPage(
            self.root,
            back_callback=self.show_home,
            level1_callback=self.show_level1_choice
        )
        self.levels_page.pack(fill="both", expand=True)

    # =====================================
    # 🎉 NIVEAU 1 : "QUESTIONS ENTRE AMIS"
    # =====================================
    def show_level1_choice(self):
        """Petit écran de choix : créer une salle, ou en rejoindre une."""
        self.clear_page()
        frame = ctk.CTkFrame(self.root, fg_color="#121620")
        frame.pack(fill="both", expand=True)

        center = ctk.CTkFrame(frame, fg_color="transparent")
        center.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(center, text="🎉", font=("Arial", 40)).pack(pady=(0, 10))
        ctk.CTkLabel(
            center, text="QUESTIONS ENTRE AMIS", font=("Arial", 18, "bold"), text_color="#FFFFFF"
        ).pack(pady=(0, 25))

        ctk.CTkButton(
            center, text="➕ CRÉER UNE SALLE", font=("Arial", 13, "bold"), width=280, height=45,
            fg_color="#8A2BE2", hover_color="#6A1B9A", command=self.show_party_create
        ).pack(pady=8)

        ctk.CTkButton(
            center, text="🔑 REJOINDRE AVEC UN CODE", font=("Arial", 13, "bold"), width=280, height=45,
            fg_color="#008080", hover_color="#004D4D", command=self.show_party_join
        ).pack(pady=8)

        ctk.CTkButton(
            center, text="← Retour", width=280, fg_color="#2B2D42", hover_color="#3A3D52",
            command=self.show_multiplayer
        ).pack(pady=(20, 0))

    def show_party_create(self):
        self.clear_page()
        self.party_create_page = PartyCreatePage(
            self.root,
            create_callback=self.verify_and_create_party_room,
            back_callback=self.show_level1_choice
        )
        self.party_create_page.pack(fill="both", expand=True)

    def verify_and_create_party_room(self, full_name, subject, nb_joueurs):
        self.show_connecting_screen("📡 Création de la salle...")
        pin_genere = f"{__import__('random').randint(100, 999)}{__import__('random').randint(100, 999)}"

        def on_response(data):
            if data.get('success'):
                confirmed_pin = data.get('pin', pin_genere)
                self.current_student_name = full_name
                self.current_active_pin = confirmed_pin
                self.party_max_players = nb_joueurs
                self.root.after(0, lambda: self.show_party_lobby(confirmed_pin, full_name, is_host=True))
            else:
                message = data.get('message', 'Erreur de connexion.')
                self.root.after(0, lambda: self._show_party_create_error(message))

        def do_connect():
            self.network.create_party_room(pin_genere, full_name, subject, nb_joueurs, response_callback=on_response)

        threading.Thread(target=do_connect, daemon=True).start()

    def _show_party_create_error(self, message):
        self.show_party_create()
        if hasattr(self, 'party_create_page'):
            self.party_create_page.show_error(message)

    def show_party_join(self):
        self.clear_page()
        self.party_join_page = JoinRoomPage(
            self.root,
            verify_join_callback=self.verify_and_join_party_room,
            back_callback=self.show_level1_choice,
            demander_sujet=True
        )
        self.party_join_page.pack(fill="both", expand=True)

    def verify_and_join_party_room(self, pin_code, full_name, subject):
        clean_pin = pin_code.replace(" ", "")
        self.show_connecting_screen("🔍 Recherche de la salle...")

        def on_response(data):
            if data.get('success'):
                self.current_student_name = full_name
                self.current_active_pin = clean_pin
                self.root.after(0, lambda: self.show_party_lobby(clean_pin, full_name, is_host=False))
            else:
                message = data.get('message', 'Erreur de connexion.')
                self.root.after(0, lambda: self._show_party_join_error(message))

        def do_connect():
            self.network.join_party_room(clean_pin, full_name, subject, on_response)

        threading.Thread(target=do_connect, daemon=True).start()

    def _show_party_join_error(self, message):
        self.show_party_join()
        if hasattr(self, 'party_join_page'):
            self.party_join_page.show_error(message)

    def show_party_lobby(self, pin, full_name, is_host):
        """Salle d'attente du mode party. Affiche le compte à rebours auto-start
        et démarre le jeu dès que le serveur envoie 'party_started'."""
        self.clear_page()
        self.current_active_pin = pin

        self.party_lobby_page = QuizLobbyPage(
            self.root,
            quiz_title="Questions entre amis",
            max_players=getattr(self, 'party_max_players', 0),
            start_quiz_callback=(self.host_start_party_quiz if is_host else None),
            back_callback=self.show_home,
            is_host=is_host
        )
        self.party_lobby_page.game_pin = pin
        if hasattr(self.party_lobby_page, 'pin_code'):
            formatted_pin = f"{pin[:3]} {pin[3:]}" if len(pin) == 6 else pin
            self.party_lobby_page.pin_code.configure(text=formatted_pin)
        self.party_lobby_page.set_players([full_name])
        self.party_lobby_page.pack(fill="both", expand=True)

        def on_player_joined(data):
            if isinstance(data, dict) and data.get('players'):
                players_list = data['players']
                self.root.after(0, lambda p=players_list: self.party_lobby_page.set_players(p))

        def on_auto_start_scheduled(data):
            secondes = data.get('seconds', 30) if isinstance(data, dict) else 30
            self.root.after(0, lambda: self._demarrer_compte_a_rebours_party(secondes))

        def on_party_started(data):
            self.root.after(0, lambda: self.start_party_game(pin, full_name))

        self.network.on_player_joined_callback = on_player_joined
        self.network.on_party_auto_start_scheduled_callback = on_auto_start_scheduled
        self.network.on_party_force_start_callback = lambda data=None: None
        self.network.on_party_started_callback = on_party_started

    def _demarrer_compte_a_rebours_party(self, secondes_restantes):
        if not hasattr(self, 'party_lobby_page') or not self.party_lobby_page.winfo_exists():
            return
        self.party_lobby_page.set_auto_start_countdown(secondes_restantes)
        if secondes_restantes > 0:
            self.root.after(1000, lambda: self._demarrer_compte_a_rebours_party(secondes_restantes - 1))

    def host_start_party_quiz(self):
        """L'hôte clique sur 'Lancer la partie' — démarrage manuel immédiat."""
        clean_pin = getattr(self, 'current_active_pin', None)
        if clean_pin:
            self.network.start_party_quiz(clean_pin)

    def start_party_game(self, pin, full_name):
        """Bascule vers l'écran de jeu du mode party dès que le serveur confirme le lancement."""
        self.clear_page()
        stop_music()

        self.party_game_page = PartyGamePage(
            self.root,
            network_controller=self.network,
            nom_joueur=full_name,
            pin=pin,
            home_callback=self.show_home
        )
        self.party_game_page.pack(fill="both", expand=True)

        self.network.on_party_question_callback = lambda data: self.root.after(
            0, lambda d=data: self.party_game_page.charger_question(d)
        )
        self.network.on_leaderboard_update_callback = lambda data: self.root.after(
            0, lambda d=data: self.party_game_page.mettre_a_jour_classement(d)
        )
        self.network.on_quiz_ended_callback = lambda data=None: self.root.after(
            0, lambda d=data: self.party_game_page.afficher_fin_partie(d)
        )
        self.network.on_party_error_callback = lambda data=None: self.root.after(
            0, lambda d=data: messagebox.showerror("Erreur", d.get('message', 'Erreur inconnue.') if d else 'Erreur inconnue.')
        )

    # =====================================
    # ÉCRAN DE CHARGEMENT RÉUTILISABLE
    # =====================================
    def show_connecting_screen(self, initial_message="Connexion en cours..."):
        """Écran de chargement générique pendant une opération réseau (rejoindre / créer une salle)."""
        self.clear_page()

        frame = ctk.CTkFrame(self.root, fg_color="#121620")
        frame.pack(fill="both", expand=True)

        center = ctk.CTkFrame(frame, fg_color="transparent")
        center.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(center, text="🔄", font=("Arial", 40)).pack(pady=(0, 15))
        self.connecting_status_lbl = ctk.CTkLabel(
            center, text=initial_message, font=("Arial", 15, "bold"), text_color="#4CC9F0"
        )
        self.connecting_status_lbl.pack()

        self.connecting_page_frame = frame
        self._connecting_messages_cycle(0)

    def _connecting_messages_cycle(self, index):
        """Fait défiler des messages rassurants pendant l'attente (peut prendre jusqu'à 60s)."""
        if not hasattr(self, 'connecting_page_frame') or not self.connecting_page_frame.winfo_exists():
            return
        messages = [
            "🔍 Recherche de la salle...",
            "📡 Connexion au serveur...",
            "⏳ Réveil du serveur, ça peut prendre un instant...",
            "🔐 Vérification en cours...",
        ]
        self.connecting_status_lbl.configure(text=messages[index % len(messages)])
        self.root.after(2500, lambda: self._connecting_messages_cycle(index + 1))

    # =====================================
    # 🔑 REJOINDRE UN QUIZ (ÉLÈVE)
    # =====================================
    def show_join_room(self):
        """Vérifie la connexion avant d'afficher le formulaire pour rejoindre un salon."""
        self._requires_login(self._show_join_room_actual)

    def _show_join_room_actual(self):
        """Affiche la page complète pour entrer le nom, prénom et le code PIN"""
        self.clear_page()
        self.join_page = JoinRoomPage(
            self.root,
            verify_join_callback=self.verify_and_join_room,
            back_callback=self.show_home
        )
        self.join_page.pack(fill="both", expand=True)

    def verify_and_join_room(self, pin_code, full_name):
        """Lance la connexion en arrière-plan avec un écran de chargement, sans jamais geler l'interface."""
        clean_pin = pin_code.replace(" ", "")
        self.show_connecting_screen("🔍 Recherche de la salle...")

        def on_response(data):
            if data.get('success'):
                self.current_student_name = full_name
                self.current_quiz_title = data.get('title') or data.get('quiz_title', 'Quiz en direct')
                initial_players = data.get('players') or [full_name]
                self.current_teacher_name = (
                    data.get('teacher_name') or
                    data.get('nom_prof') or
                    data.get('professeur') or
                    "Professeur"
                )
                self.current_active_pin = clean_pin

                self.root.after(0, lambda p=initial_players: self.show_student_lobby(
                    clean_pin, self.current_quiz_title, full_name, initial_players=p
                ))
            else:
                message = data.get('message', 'Erreur de connexion.')
                self.root.after(0, lambda m=message: self.show_join_room_with_error(m))

        def do_connect():
            self.network.join_room(clean_pin, full_name, on_response)

        threading.Thread(target=do_connect, daemon=True).start()

    def show_join_room_with_error(self, message):
        """Revient au formulaire de connexion avec un message d'erreur affiché."""
        self.show_join_room()
        if hasattr(self, 'join_page'):
            self.join_page.show_error(message)

    def show_student_lobby(self, pin, title, full_name, initial_players=None):
        """Affiche la salle d'attente pour l'élève et s'abonne au lancement du jeu"""
        self.clear_page()
        self.current_active_pin = pin  # 👈 Sécurité PIN supplémentaire
        self.lobby_page = QuizLobbyPage(
            self.root,
            quiz_title=title,
            max_players=0,
            start_quiz_callback=None,
            back_callback=self.show_home,
            is_host=False
        )
        
        self.lobby_page.game_pin = pin
        if hasattr(self.lobby_page, 'pin_code'):
            formatted_pin = f"{pin[:3]} {pin[3:]}" if len(pin) == 6 else pin
            self.lobby_page.pin_code.configure(text=formatted_pin)

        self.lobby_page.set_players(initial_players or [full_name])

        # Mettre à jour la liste quand un joueur rejoint
        def on_player_joined(data):
            if isinstance(data, dict) and data.get('players'):
                players_list = data['players']
                self.root.after(0, lambda p=players_list: self.lobby_page.set_players(p))

        self.network.on_player_joined_callback = on_player_joined

        # 🚀 RÉCEPTION DE L'ÉVÉNEMENT DE DÉMARRAGE DU QUIZ
        def handle_start_quiz(data=None):
            print("🟢 ÉVÉNEMENT REÇU CÔTÉ ÉLÈVE : Lancement du Quiz !", data)
            
            questions = []
            if isinstance(data, dict):
                questions = data.get('questions', [])
                if 'title' in data and data['title']:
                    self.current_quiz_title = data['title']
                
                # Récupération sécurisée du nom du professeur
                prof = (
                    data.get('teacher_name') or 
                    data.get('nom_prof') or 
                    data.get('professeur')
                )
                if prof:
                    self.current_teacher_name = prof

            if questions:
                self.current_quiz = questions

            # Basculer l'affichage dans le thread Tkinter principal
            def update_ui():
                if hasattr(self, 'lobby_page') and self.lobby_page:
                    self.lobby_page.pack_forget()
                self.start_student_game(questions_list=questions)

            self.root.after(0, update_ui)

        # Enregistrement propre via la fonction de rappel de NetworkClient
        self.network.on_quiz_started_callback = handle_start_quiz

        self.lobby_page.pack(fill="both", expand=True)

    def start_student_game(self, questions_list=None):
        """Démarre la session de jeu MANUELLE pour l'élève"""
        print("▶️ Transition vers PlayQuizManuelPage pour l'élève...")
        self.clear_page()
        stop_music()

        if questions_list:
            self.current_quiz = questions_list

        # 🎯 Transmission du network_controller + métadonnées + PIN
        self.play = PlayQuizManuelPage(
            self.root,
            network_controller=self.network,
            titre_quiz=self.current_quiz_title,
            nom_prof=self.current_teacher_name,
            nom_eleve=self.current_student_name,
            home_callback=self.show_home,
            pin=getattr(self, 'current_active_pin', None)  # 👈 Transmission du PIN
        )
        self.play.pack(fill="both", expand=True)

        # 🚀 Charger immédiatement la première question transmise
        if self.current_quiz and len(self.current_quiz) > 0:
            first_q = self.current_quiz[0]
            first_q['teacher_name'] = self.current_teacher_name
            self.play.load_network_question(
                question_data=first_q,
                current_index=1,
                total_questions=len(self.current_quiz)
            )

        def handle_change_question(data):
            new_index = data.get('question_index', 0)
            print(f"📩 Ordre serveur reçu côté élève : passer à l'index {new_index}")
            self.root.after(0, lambda: self.passer_question_suivante_eleve(new_index))

        # Liaison avec le réseau WebSocket
        self.network.on_change_question_callback = handle_change_question
        self.network.on_leaderboard_update_callback = lambda data: self.root.after(0, lambda d=data: self.play.mettre_a_jour_classement(d))
        self.network.on_quiz_ended_callback = lambda data=None: self.root.after(0, lambda d=data: self.play.recevoir_classement_final(d))
        self.network.on_quiz_cancelled_callback = lambda data=None: self.root.after(0, self.play.afficher_quiz_annule)

    def passer_question_suivante_eleve(self, index):
        """Passe la question de l'élève à l'index demandé par le serveur"""
        if hasattr(self, 'play') and isinstance(self.play, PlayQuizManuelPage):
            if self.current_quiz and 0 <= index < len(self.current_quiz):
                question_data = self.current_quiz[index]
                question_data['teacher_name'] = self.current_teacher_name
                self.play.load_network_question(
                    question_data=question_data,
                    current_index=index + 1,
                    total_questions=len(self.current_quiz)
                )

    def show_settings(self):
        self._requires_login(self._show_settings_actual)

    def _show_settings_actual(self):
        self.clear_page()
        self.settings_page = SettingsPage(
            self.root,
            back_callback=self.show_home,
            account_callback=self.show_account,
        )
        self.settings_page.pack(fill="both", expand=True)

    def show_account(self):
        self._requires_login(self._show_account_actual)

    def _show_account_actual(self):
        self.clear_page()
        self.account_page = AccountPage(
            self.root,
            back_callback=self.show_home,
            on_logout=self._apres_deconnexion,
        )
        self.account_page.pack(fill="both", expand=True)

    def _apres_deconnexion(self):
        self.show_home()

    # =====================================
    # 📁 MES QUIZ & BIBLIOTHÈQUE HUB
    # =====================================
    def show_my_quizzes(self):
        """Vérifie la connexion avant d'afficher la bibliothèque (privée au compte)."""
        self._requires_login(self._show_my_quizzes_actual)

    def _show_my_quizzes_actual(self):
        """Affiche proprement la page du Hub Bibliothèque intégrée dans l'application."""
        self.clear_page()
        self.library_hub = LibraryHubWindow(master=self.root, app_controller=self)
        self.library_hub.pack(fill="both", expand=True)

    def start_game(self, quiz_data):
        """Lancement d'un quiz depuis la bibliothèque (ouvre le salon de jeu)."""
        if not quiz_data:
            return
        title = quiz_data.get("title", "Quiz")
        questions = quiz_data.get("questions", [])
        teacher_name = quiz_data.get("teacher_name", self.current_teacher_name)
        
        self.open_quiz_lobby(title, questions, teacher_name=teacher_name)

    def open_editor(self, quiz_data=None):
        """Ouvre l'éditeur de quiz manuel avec les données fournies."""
        self.show_create_manual()
        
        if quiz_data and hasattr(self, 'manual_page'):
            # On laisse 100ms à Tkinter pour construire entièrement les widgets de manual_page
            def apply_quiz_data():
                if hasattr(self.manual_page, 'load_quiz'):
                    self.manual_page.load_quiz(quiz_data)
                elif hasattr(self.manual_page, 'load_quiz_data'):
                    self.manual_page.load_quiz_data(quiz_data)
                elif hasattr(self.manual_page, 'set_quiz'):
                    self.manual_page.set_quiz(quiz_data)

            self.root.after(100, apply_quiz_data)

    def open_quiz_lobby(self, quiz_title, questions, teacher_name="Professeur"):
        """Lance la connexion en arrière-plan, puis ouvre la vraie salle une fois confirmée par le serveur."""
        self.current_quiz = questions
        self.current_quiz_title = quiz_title
        self.current_teacher_name = teacher_name

        self.show_connecting_screen("📡 Création de la salle...")

        pin_genere = f"{__import__('random').randint(100, 999)}{__import__('random').randint(100, 999)}"

        def on_room_created(data):
            if data.get('success'):
                confirmed_pin = data.get('pin', pin_genere)
                self.root.after(0, lambda: self._show_confirmed_host_lobby(confirmed_pin))
            else:
                self.root.after(0, self._show_lobby_creation_error)

        def do_connect():
            self.network.create_room(pin_genere, quiz_title, teacher_name=teacher_name, response_callback=on_room_created)

        threading.Thread(target=do_connect, daemon=True).start()

    def _show_confirmed_host_lobby(self, confirmed_pin):
        """Affiche la vraie salle d'attente une fois le serveur a réellement confirmé sa création."""
        self.clear_page()
        self.current_active_pin = confirmed_pin

        self.lobby_page = QuizLobbyPage(
            self.root,
            quiz_title=self.current_quiz_title,
            max_players=0,
            start_quiz_callback=self.host_start_quiz_callback,
            back_callback=self.show_my_quizzes,
            is_host=True
        )

        # Remplace le PIN généré en interne par QuizLobbyPage par le VRAI PIN confirmé par le serveur
        self.lobby_page.game_pin = confirmed_pin
        if hasattr(self.lobby_page, 'pin_code'):
            formatted_pin = f"{confirmed_pin[:3]} {confirmed_pin[3:]}" if len(confirmed_pin) == 6 else confirmed_pin
            self.lobby_page.pin_code.configure(text=formatted_pin)

        def on_player_joined_remote(data):
            if isinstance(data, dict) and data.get('players'):
                players_list = data['players']
                self.root.after(0, lambda p=players_list: self.lobby_page.set_players(p))

        self.network.on_player_joined_callback = on_player_joined_remote
        self.lobby_page.pack(fill="both", expand=True)

    def _show_lobby_creation_error(self):
        """Affiche un message d'erreur clair si la salle n'a jamais pu être créée."""
        self.clear_page()
        frame = ctk.CTkFrame(self.root, fg_color="#121620")
        frame.pack(fill="both", expand=True)

        center = ctk.CTkFrame(frame, fg_color="transparent")
        center.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(center, text="⚠️", font=("Arial", 40)).pack(pady=(0, 15))
        ctk.CTkLabel(
            center, text="Impossible de contacter le serveur.\nVérifie ta connexion et réessaie.",
            font=("Arial", 14, "bold"), text_color="#FF5252", justify="center"
        ).pack(pady=(0, 20))

        ctk.CTkButton(
            center, text="🔄 Réessayer", fg_color="#1F6AA5", hover_color="#144870",
            command=lambda: self.open_quiz_lobby(self.current_quiz_title, self.current_quiz, self.current_teacher_name)
        ).pack(pady=(0, 10))

        ctk.CTkButton(
            center, text="🏠 Retour à l'accueil", fg_color="#37474F", hover_color="#263238",
            command=self.show_home
        ).pack()

    def host_start_quiz_callback(self):
        """Fonction appelée lorsque l'hôte clique sur LANCER LE QUIZ"""
        clean_pin = getattr(self, 'current_active_pin', None)
        print(f"🔴 L'hôte lance le quiz pour la salle PIN: {clean_pin}")

        if clean_pin:
            self.network.start_quiz(
                clean_pin, 
                self.current_quiz, 
                self.current_teacher_name,
                self.current_quiz_title
            )

        if hasattr(self, 'lobby_page') and self.lobby_page:
            self.lobby_page.pack_forget()

        self.host_current_question_index = 0
        self.latest_host_leaderboard = []
        self.quiz_finalise = False

        # Calcule le total de points possible du quiz (ex: 40, 70...)
        try:
            self.host_total_points = sum(
                int("".join(ch for ch in str(q.get("points", 10)) if ch.isdigit()) or 10)
                for q in (self.current_quiz or [])
            )
        except Exception:
            self.host_total_points = 0

        # Écoute le classement en direct dès le lancement (utile pendant ET après le quiz)
        if hasattr(self.network, 'sio') and self.network.sio:
            @self.network.sio.on('leaderboard_update')
            def on_leaderboard_update(data):
                players_raw = data.get('players', []) if isinstance(data, dict) else data
                self.latest_host_leaderboard = players_raw
                if hasattr(self, 'teacher_dashboard') and self.teacher_dashboard.winfo_exists():
                    self.root.after(0, lambda f=players_raw: self.teacher_dashboard.update_dashboard(f))

            # 🔗 Vrai signal de synchronisation : se déclenche dès que TOUS les élèves
            # ont fini la dernière question — en même temps que leur propre page résultat.
            @self.network.sio.on('quiz_ended')
            def on_quiz_ended_host(data):
                self.root.after(4500, self._finish_quiz_now)

        self.show_host_live_view()

    def show_host_live_view(self):
        """Vue spectateur du prof pendant le quiz : voit la question en direct, contrôle le rythme."""
        self.clear_page()
        stop_music()

        container = ctk.CTkFrame(self.root, fg_color="#0F111A")
        container.pack(fill="both", expand=True, padx=20, pady=20)

        control_bar = ctk.CTkFrame(container, fg_color="#1E222D", corner_radius=12)
        control_bar.pack(fill="x", pady=(0, 15), ipady=8)

        ctk.CTkLabel(
            control_bar, text=f"👁️ MODE SPECTATEUR — PIN : {getattr(self, 'current_active_pin', '')}",
            font=("Arial", 14, "bold"), text_color="#FFD700"
        ).pack(side="left", padx=15)

        self.host_timer_lbl = ctk.CTkLabel(control_bar, text="⏱ --s", font=("Arial", 14, "bold"), text_color="#4CC9F0")
        self.host_timer_lbl.pack(side="left", padx=15)

        self.next_q_btn = ctk.CTkButton(
            control_bar, text="Question Suivante ➔", fg_color="#2FA572", hover_color="#1E6B49",
            font=("Arial", 13, "bold"), state="disabled", command=self.host_click_next_question
        )
        self.next_q_btn.pack(side="left", padx=15)

        ctk.CTkButton(
            control_bar, text="🛑 Terminer le quiz", fg_color="#D32F2F", hover_color="#9A0007",
            font=("Arial", 13, "bold"), command=self.host_cancel_quiz
        ).pack(side="right", padx=15)

        self.host_question_card = ctk.CTkFrame(container, fg_color="#181B26", corner_radius=16, border_width=1, border_color="#242838")
        self.host_question_card.pack(fill="both", expand=True, ipady=20, ipadx=20)

        self.host_progress_lbl = ctk.CTkLabel(self.host_question_card, text="", font=("Arial", 12, "bold"), text_color="#A0AABF")
        self.host_progress_lbl.pack(pady=(15, 5))

        self.host_question_lbl = ctk.CTkLabel(
            self.host_question_card, text="", font=("Arial", 20, "bold"), text_color="#FFFFFF", wraplength=800
        )
        self.host_question_lbl.pack(pady=(10, 25))

        self.host_options_frame = ctk.CTkFrame(self.host_question_card, fg_color="transparent")
        self.host_options_frame.pack(fill="x", padx=40)
        self.host_options_frame.columnconfigure(0, weight=1)
        self.host_options_frame.columnconfigure(1, weight=1)

        self.host_option_buttons = {}
        for i, lettre in enumerate(["A", "B", "C", "D"]):
            btn = ctk.CTkButton(
                self.host_options_frame, text="", font=("Arial", 13, "bold"),
                fg_color="#242838", hover_color="#242838", state="disabled",
                height=50, corner_radius=10, text_color_disabled="#FFFFFF"
            )
            btn.grid(row=i // 2, column=i % 2, padx=8, pady=8, sticky="ew")
            self.host_option_buttons[lettre] = btn

        self.render_host_question()
        self.start_host_timer()

    def render_host_question(self):
        """Affiche la question actuelle en lecture seule pour le prof — SANS révéler la bonne réponse (utile en projection)."""
        if not self.current_quiz or self.host_current_question_index >= len(self.current_quiz):
            return

        q = self.current_quiz[self.host_current_question_index]
        total = len(self.current_quiz)

        self.host_progress_lbl.configure(text=f"Question {self.host_current_question_index + 1} / {total}")
        self.host_question_lbl.configure(text=q.get("question", ""))

        for lettre, btn in self.host_option_buttons.items():
            texte = q.get(lettre, "")
            btn.configure(text=f"{lettre}.  {texte}", fg_color="#242838", text_color_disabled="#FFFFFF")

    def host_cancel_quiz(self):
        """Le prof met fin au quiz prématurément — tous les élèves sont notifiés."""
        clean_pin = getattr(self, 'current_active_pin', None)
        if clean_pin:
            self.network.cancel_quiz(clean_pin)
        self.show_home()

    def host_click_next_question(self):
        """Appelé lorsque le professeur clique sur le bouton Question Suivante (actif seulement chrono écoulé)"""
        clean_pin = getattr(self, 'current_active_pin', None)
        if clean_pin:
            print(f"⏭️ Le professeur demande le passage à la question suivante pour la salle [{clean_pin}]")
            self.network.send_next_question(clean_pin)

        self.host_current_question_index += 1

        if self.current_quiz and self.host_current_question_index < len(self.current_quiz):
            self.render_host_question()
            self.start_host_timer()

    def show_host_dashboard(self):
        """Affiche le Tableau de bord final complet (classement, statistiques, export)"""
        self.clear_page()
        stop_music()

        container = ctk.CTkFrame(self.root, fg_color="#121620")
        container.pack(fill="both", expand=True, padx=20, pady=20)

        header = ctk.CTkFrame(container, fg_color="transparent")
        header.pack(fill="x", pady=(0, 5))

        ctk.CTkLabel(
            header, text=f"🏁 QUIZ TERMINÉ — PIN : {getattr(self, 'current_active_pin', '')}",
            font=("Arial", 16, "bold"), text_color="#FFD700"
        ).pack(side="left")

        ctk.CTkButton(
            header, text="Terminer la session", fg_color="#D32F2F", hover_color="#9A0007",
            command=self.show_home
        ).pack(side="right")

        meta_bar = ctk.CTkFrame(container, fg_color="#1E222D", corner_radius=10)
        meta_bar.pack(fill="x", pady=(10, 15), ipady=8)

        meta_text = (
            f"📘 {self.current_quiz_title}     "
            f"👤 {self.current_teacher_name}     "
            f"📅 {datetime.now().strftime('%d %b %Y')}     "
            f"🏆 Quiz noté sur {getattr(self, 'host_total_points', 0)} points"
        )
        ctk.CTkLabel(meta_bar, text=meta_text, font=("Arial", 12, "bold"), text_color="#A0AABF").pack(padx=15)

        self.teacher_dashboard = TeacherFullDashboard(container)
        self.teacher_dashboard.pack(fill="both", expand=True)
        self.teacher_dashboard.update_dashboard(getattr(self, 'latest_host_leaderboard', []))

        export_bar = ctk.CTkFrame(container, fg_color="transparent")
        export_bar.pack(fill="x", pady=(15, 0))

        ctk.CTkButton(
            export_bar, text="📥 Exporter en PDF", font=("Arial", 12, "bold"),
            fg_color="#1F6AA5", hover_color="#144870", command=self.export_results_pdf
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            export_bar, text="📊 Exporter en Excel", font=("Arial", 12, "bold"),
            fg_color="#2E7D32", hover_color="#1E4620", command=self.export_results_excel
        ).pack(side="left")

    def export_results_pdf(self):
        """Exporte le classement actuel en PDF"""
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("Module manquant", "Installe d'abord : pip install fpdf2")
            return

        data = getattr(self, 'latest_host_leaderboard', [])
        if not data:
            messagebox.showinfo("Export PDF", "Aucun résultat à exporter pour l'instant.")
            return

        chemin = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Fichier PDF", "*.pdf")],
            initialfile=f"{self.current_quiz_title}_resultats.pdf"
        )
        if not chemin:
            return

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, f"Résultats : {self.current_quiz_title}", ln=True)
        pdf.set_font("Arial", "", 11)
        pdf.cell(0, 8, f"Professeur : {self.current_teacher_name}", ln=True)
        pdf.cell(0, 8, f"Date : {datetime.now().strftime('%d %b %Y')}", ln=True)
        pdf.cell(0, 8, f"Barème : {getattr(self, 'host_total_points', 0)} points", ln=True)
        pdf.ln(6)

        pdf.set_font("Arial", "B", 11)
        pdf.cell(20, 8, "Rang", border=1)
        pdf.cell(70, 8, "Nom", border=1)
        pdf.cell(30, 8, "Points", border=1)
        pdf.cell(30, 8, "Combo", border=1)
        pdf.cell(30, 8, "Temps Moy.", border=1, ln=True)

        pdf.set_font("Arial", "", 11)
        for p in data:
            pdf.cell(20, 8, str(p.get("rank", "")), border=1)
            pdf.cell(70, 8, str(p.get("name", "")), border=1)
            pdf.cell(30, 8, str(p.get("score", 0)), border=1)
            pdf.cell(30, 8, f"x{p.get('combo', 0)}", border=1)
            pdf.cell(30, 8, f"{p.get('avg_time', 0)}s", border=1, ln=True)

        pdf.output(chemin)
        messagebox.showinfo("Export PDF", "Fichier PDF exporté avec succès !")

    def export_results_excel(self):
        """Exporte le classement actuel en Excel (.xlsx)"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Module manquant", "Installe d'abord : pip install openpyxl")
            return

        data = getattr(self, 'latest_host_leaderboard', [])
        if not data:
            messagebox.showinfo("Export Excel", "Aucun résultat à exporter pour l'instant.")
            return

        chemin = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
            initialfile=f"{self.current_quiz_title}_resultats.xlsx"
        )
        if not chemin:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"

        ws.append([f"Quiz : {self.current_quiz_title}"])
        ws.append([f"Professeur : {self.current_teacher_name}"])
        ws.append([f"Date : {datetime.now().strftime('%d %b %Y')}"])
        ws.append([f"Barème : {getattr(self, 'host_total_points', 0)} points"])
        ws.append([])
        ws.append(["Rang", "Nom", "Points", "Combo", "Temps Moyen (s)"])

        for p in data:
            ws.append([
                p.get("rank", ""), p.get("name", ""), p.get("score", 0),
                p.get("combo", 0), p.get("avg_time", 0)
            ])

        wb.save(chemin)
        messagebox.showinfo("Export Excel", "Fichier Excel exporté avec succès !")

    def start_host_timer(self):
        """Démarre le chrono du prof pour la question en cours et verrouille le bouton Suivant."""
        if not self.current_quiz or self.host_current_question_index >= len(self.current_quiz):
            return

        question = self.current_quiz[self.host_current_question_index]
        time_str = question.get("time", "20s")
        try:
            seconds = int("".join(ch for ch in str(time_str) if ch.isdigit()) or 20)
        except ValueError:
            seconds = 20

        self.host_remaining_time = seconds
        self.next_q_btn.configure(state="disabled")
        self._host_timer_tick()

    def _host_timer_tick(self):
        """Décompte du chrono prof, tick par seconde."""
        if not hasattr(self, 'next_q_btn'):
            return

        if self.host_remaining_time <= 0:
            self.next_q_btn.configure(state="normal", text="Question Suivante ➔")
            self.host_timer_lbl.configure(text="✅ Prêt")

            # Sur la DERNIÈRE question : on attend UNIQUEMENT le vrai signal
            # "tout le monde a répondu" (voir on_quiz_ended_host).
            est_derniere = self.current_quiz and self.host_current_question_index >= len(self.current_quiz) - 1
            if est_derniere:
                self.next_q_btn.configure(state="disabled", text="⏳ En attente des élèves...")
            return

        self.host_timer_lbl.configure(text=f"⏱ {self.host_remaining_time}s")
        self.next_q_btn.configure(text=f"⏳ Patiente... ({self.host_remaining_time}s)")
        self.host_remaining_time -= 1
        self.root.after(1000, self._host_timer_tick)

    def _finish_quiz_now(self):
        """Déclenché par le vrai signal réseau 'quiz_ended' — synchronisé avec les élèves."""
        if getattr(self, 'quiz_finalise', False):
            return
        self.quiz_finalise = True
        self.show_host_dashboard()

    def _fallback_finish_quiz(self):
        """Filet de sécurité si la synchro automatique n'arrive jamais."""
        if getattr(self, 'quiz_finalise', False):
            return
        clean_pin = getattr(self, 'current_active_pin', None)
        if clean_pin:
            self.network.send_next_question(clean_pin)

    # =====================================
    # 📝 Création manuelle
    # =====================================
    def show_manual_choice(self):
        """Affiche le choix Créer / Rejoindre pour le Mode Classe."""
        self.clear_page()
        self.manual_choice_page = ManualQuizChoicePage(
            self.root,
            create_callback=self.show_create_manual,
            join_callback=self.show_join_room,
            back_callback=self.show_home,
        )
        self.manual_choice_page.pack(fill="both", expand=True)

    def show_create_manual(self):
        self.clear_page()

        if ManualQuizPage:
            self.manual_page = ManualQuizPage(
                self.root,
                on_quiz_created=self.on_manual_quiz_created,
                back_callback=self.show_home
            )
            self.manual_page.pack(fill="both", expand=True)
        else:
            print("Erreur : Le fichier ui/manual_quiz.py est introuvable.")

    def on_manual_quiz_created(self, quiz_title="Nouveau Quiz", teacher_name="Professeur", quiz_data=None, **kwargs):
        """Gère la création du quiz manuel en capturant le titre, le nom du professeur et les questions."""
        if quiz_data:
            save_quiz(quiz_title, quiz_data, teacher_name=teacher_name)
            print(f"Quiz '{quiz_title}' créé par {teacher_name} enregistré localement avec succès !")
        
        self.show_my_quizzes()

    # =====================================
    # Générateur TMY (IA)
    # =====================================
    def show_tmy_generator(self):
        """Vérifie la connexion avant de lancer le générateur de quiz par IA."""
        self._requires_login(self._show_tmy_generator_actual)

    def _show_tmy_generator_actual(self):
        self.clear_page()

        self.tmy_generator = TMYGeneratorPage(
            self.root,
            self.start_quiz,
            self.show_home
        )

        self.tmy_generator.pack(fill="both", expand=True)

    def start_quiz(self, sujet, nombre, niveau):
        self.current_quiz_settings = {
            "sujet": sujet,
            "nombre": nombre,
            "niveau": niveau
        }

        print("Création du quiz via TMY AI...")
        self.show_quiz_loading()

        threading.Thread(
            target=self.generate_quiz_background,
            daemon=True
        ).start()

    def generate_quiz_background(self):
        succes, message, donnees = session.generer_quiz(
            self.current_quiz_settings["sujet"],
            self.current_quiz_settings["nombre"],
            self.current_quiz_settings["niveau"],
        )

        if succes:
            self.current_quiz = donnees.get("questions", [])
            self.root.after(0, self.show_play_quiz)
        else:
            print("Erreur TMY :", message)
            self.root.after(0, lambda: self.show_tmy_generator())

    # =====================================
    # Écran de chargement Quiz
    # =====================================
    def show_quiz_loading(self):
        self.clear_page()

        self.quiz_loading = QuizLoadingPage(self.root)
        self.quiz_loading.pack(fill="both", expand=True)

    # =====================================
    # Affichage du Quiz en cours (MODE SOLO)
    # =====================================
    def show_play_quiz(self):
        self.clear_page()
        stop_music()

        # On construit un vrai dict (sujet + niveau + questions) au lieu de passer
        # juste la liste de questions — sinon PlayQuizPage ne peut pas connaître
        # le titre ni le niveau et retombe sur "Quiz Solo" par défaut.
        quiz_pour_page = {
            "sujet": self.current_quiz_settings.get("sujet", "Quiz Solo"),
            "niveau": self.current_quiz_settings.get("niveau", ""),
            "questions": self.current_quiz,
        }

        self.play = PlayQuizPage(
            self.root,
            quiz_pour_page,
            self.show_result,
            cancel_callback=self.show_home,
        )

        self.play.pack(fill="both", expand=True)

    # =====================================
    # Écran des Résultats Solo
    # =====================================
    def show_result(
        self,
        score,
        total,
        total_xp,
        max_combo,
        average_time=0,
        quiz_data=None
    ):
        print("========== RESULT ==========")
        print("Score :", score)
        print("Total :", total)
        print("XP :", total_xp)
        print("Combo :", max_combo)
        print("============================")

        self.clear_page()

        self.root.after(2500, resume_music)

        quiz_data = quiz_data or {}
        titre_quiz = quiz_data.get("sujet") or quiz_data.get("quiz_title") or "Quiz"
        niveau_quiz = quiz_data.get("niveau", "")

        self.result = ResultPage(
            self.root,
            score,
            total,
            total_xp,
            max_combo,
            average_time,
            self.regenerate_quiz,
            self.show_home,
            quiz_title=titre_quiz,
            niveau=niveau_quiz,
        )

        self.result.pack(fill="both", expand=True)

    # =====================================
    # Régénération du Quiz
    # =====================================
    def regenerate_quiz(self):
        settings = self.current_quiz_settings
        self.show_quiz_loading()

        def regen():
            succes, message, donnees = session.generer_quiz(
                settings["sujet"],
                settings["nombre"],
                settings["niveau"],
                regeneration=True,
            )

            if succes:
                self.current_quiz = donnees.get("questions", [])
                self.root.after(0, self.show_play_quiz)
            else:
                print("Erreur régénération :", message)
                self.root.after(0, lambda: self.show_home())

        threading.Thread(target=regen, daemon=True).start()

    # =====================================
    # Nettoyage de l'écran
    # =====================================
    def clear_page(self):
        self.loading_active = False
        for widget in list(self.root.winfo_children()):
            widget.destroy()
